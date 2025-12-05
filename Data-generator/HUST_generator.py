import sys
import os
import pandas as pd
import numpy as np
import re
from scipy.io import loadmat
from numpy.lib.stride_tricks import sliding_window_view

# 获取项目根目录（论文4目录）的绝对路径
# project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
# # 将根目录添加到sys.path
# sys.path.append(project_root)
sys.path.append('/root/hy-tmp')  # 手动加入 hy-tmp 的路径

from Data.HUST_dir import hust_L1, hust_L2, hust_L3, hust_L4
from MTAGN_utils.data_preprocess import add_wgn
from MTAGN_utils.data_preprocess import standardization
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

def data_labels_shuffle(data, label1, label2):
    """
    input 二维numpy array
    :param data: [data_num, data_len]
    :param label1: [data_num,]
    :param label2: [data_num,]
    :return:
    """
    np.random.seed(2021)
    index = np.arange(len(data))
    np.random.shuffle(index)
    return data[index], label1[index], label2[index]

def resolve_single_xlsx(dir_or_file: str) -> str:
    """
    只返回唯一的 .xlsx/.xlsm 文件路径：
    - 如果传入的是文件：必须是 .xlsx/.xlsm
    - 如果传入的是目录：只在目录里找 .xlsx/.xlsm（忽略 .xls、~$ 临时文件、隐藏目录）
      必须恰好 1 个，否则报错并列出找到的项
    """
    p = Path(dir_or_file)
    if p.is_file():
        if p.suffix.lower() in (".xlsx", ".xlsm") and not p.name.startswith("~$"):
            return str(p.resolve())
        raise ValueError(f"Expect .xlsx/.xlsm file, got: {p.name}")

    if not p.is_dir():
        raise FileNotFoundError(f"Not found: {dir_or_file}")

    cands = [f for f in sorted(p.glob("*.xlsx")) + sorted(p.glob("*.xlsm"))
             if f.is_file() and not f.name.startswith("~$")]
    if len(cands) == 0:
        raise FileNotFoundError(f"No .xlsx/.xlsm in '{p}'")
    if len(cands) > 1:
        names = ", ".join(x.name for x in cands)
        raise RuntimeError(f"There are {len(cands)} .xlsx/.xlsm in '{p}': {names}. Please keep only one!")
    return str(cands[0].resolve())

def get_HUST_data(path_or_dir, col_idx: int = 4, sheet_name: str = None):
    """
    只读取 .xlsx/.xlsm（忽略 .xls）：
    - 在首列找到 'Data'（大小写不敏感，可带冒号）
    - 从下一行开始读取第 col_idx 列（默认 4 -> 第5列）
    - 若该列为空，兜底：整行提数后取第 col_idx 个
    返回 temp_data: 1D float32
    """
    p = _resolve_single_xlsx(path_or_dir)  # ← 只会返回 .xlsx/.xlsm

    wb = load_workbook(p, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    # 找到 'Data' 行
    data_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        v0 = row[0] if row and len(row) > 0 else None
        s = str(v0).strip().lower() if v0 is not None else ""
        if s in ("data", "data:"):
            data_row = i
            break
    if data_row is None:
        raise ValueError(f"未找到 'Data' 标记：{p.name}")

    # 直接取目标列
    vals = []
    c = col_idx + 1  # 1-based 列号
    for row in ws.iter_rows(min_row=data_row + 1, min_col=c, max_col=c, values_only=True):
        v = row[0]
        if v is None:
            continue
        try:
            vals.append(float(str(v).strip().replace("−","-").replace("\u2212","-")))
        except:
            pass

    # 兜底：整行提数
    if not vals:
        for row in ws.iter_rows(min_row=data_row + 1, values_only=True):
            if not row:
                continue
            parts = [str(x).strip() for x in row if x is not None and str(x).strip() != ""]
            if not parts:
                continue
            sline = " ".join(parts).replace("−","-").replace("\u2212","-")
            nums = re.findall(r"[+-]?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?", sline)
            if len(nums) > col_idx:
                try:
                    vals.append(float(nums[col_idx]))
                except:
                    pass

    if not vals:
        raise ValueError(f"未在 {p.name} 的 'Data' 之后解析到第 {col_idx+1} 列的数值")

    temp_data = np.asarray(vals, dtype=np.float32)
    return temp_data  # 保持返回名

def get_HUST_data(file_dir, col_idx: int = 4, sheet_name: str = None):
    """
    用 openpyxl 读取 .xlsx/.xlsm：
    - 在首列找到 'Data'（大小写不敏感，可带冒号）
    - 从其下一行开始读取第 col_idx 列（默认=4 -> 第5列）
    - 若第5列为空，则行级兜底：拼行->正则提数->取第 col_idx 个
    返回: temp_data (np.float32, 1D)
    """
    p = Path(file_dir)
    if p.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError(f"openpyxl 只支持 .xlsx/.xlsm，当前文件：{p.name}")

    wb = load_workbook(p, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    # 1) 找到 'Data' 行
    data_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        v0 = row[0] if row and len(row) > 0 else None
        s = str(v0).strip().lower() if v0 is not None else ""
        if s in ("data", "data:"):
            data_row = i
            break
    if data_row is None:
        raise ValueError(f"未找到 'Data' 标记：{p.name}")

    # 2) 正常路径：直接读第5列（1-based 列号 = col_idx+1）
    vals = []
    col_1based = col_idx + 1
    for row in ws.iter_rows(min_row=data_row + 1, min_col=col_1based, max_col=col_1based, values_only=True):
        v = row[0]
        if v is None:
            continue
        try:
            vals.append(float(str(v).strip().replace("−", "-").replace("\u2212", "-")))
        except Exception:
            continue

    # 3) 兜底：如果上面没拿到，逐行把整行拼成字符串再正则提取数字，取第5个
    if not vals:
        for row in ws.iter_rows(min_row=data_row + 1, values_only=True):
            if not row:
                continue
            parts = [str(x).strip() for x in row if x is not None and str(x).strip() != ""]
            if not parts:
                continue
            sline = " ".join(parts).replace("−", "-").replace("\u2212", "-")
            nums = re.findall(r"[+-]?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?", sline)
            if len(nums) > col_idx:
                try:
                    vals.append(float(nums[col_idx]))
                except Exception:
                    pass

    if not vals:
        raise ValueError(f"未在 {p.name} 的 'Data' 之后解析到第 {col_idx+1} 列的数值")

    temp_data = np.asarray(vals, dtype=np.float32)
    return temp_data


def HUST_preprocess(file_dir, data_len=2048, sample_num=450, overlap=True, overlap_step=512, shuffle=True,
                  add_noise=True, SNR=5, normalize=True, split_rate=(60, 10, 30), label1=None, label2=None):
    """

    :param file_dir:
    :param data_len:
    :param sample_num:
    :param overlap:
    :param overlap_step:
    :param shuffle:
    :param add_noise:
    :param SNR:
    :param normalize:
    :param split_rate:
    :param label1:
    :param label2:
    :return:
    """
    def data_sampling(raw_data):
        """
        对原始数据(199935个点)进行取样
        :param raw_data: [199935,]
        :return: samples: [num, data_len]
        """
        samples = np.empty((0, data_len), dtype=np.float32)
        if not overlap:  # 不重叠取样
            start = 0
            for i in range(sample_num):
                sample = raw_data[start: start + data_len]
                samples = np.vstack((samples, sample))
                start += data_len
        else:  # 重叠取样
            start = 0
            for i in range(sample_num):
                sample = raw_data[start: start + data_len]
                samples = np.vstack((samples, sample))
                start += overlap_step
        if add_noise:
            samples = add_wgn(samples, snr=SNR)
        if normalize:
            samples = standardization(samples)
        return samples

    def add_labels(x):
        """
        add fault type label and fault severity label
        :param x:
        :return: labels1-[data_len,], labels2-[data_len,]
        """
        len_x = len(x)
        labels_1 = np.ones((len_x,), dtype=np.int64) * label1
        labels_2 = np.ones((len_x,), dtype=np.int64) * label2
        return labels_1, labels_2

    def split_data(samples, labels1, labels2):
        """
        split data into train_set, valid_set and test_set
        :param samples:
        :param labels1:
        :param labels2:
        :return:
        """
        # if sum(split_rate) != sample_num:
        #     print("error, sum(split_rate != sample_num)")
        #     exit()
        end_point1 = int(split_rate[0])
        end_point2 = int(split_rate[0] + split_rate[1])
        end_point3 = sum(split_rate)
        x_train, y1_train, y2_train = samples[:end_point1], labels1[:end_point1], labels2[:end_point1]
        x_valid, y1_valid, y2_valid = samples[end_point1: end_point2], labels1[end_point1: end_point2], \
                                      labels2[end_point1: end_point2]
        x_test, y1_test, y2_test = samples[end_point2: end_point3], labels1[end_point2: end_point3], \
                                   labels2[end_point2: end_point3]
        return x_train, y1_train, y2_train, x_valid, y1_valid, y2_valid, x_test, y1_test, y2_test

    data = get_HUST_data(file_dir)
    Samples = data_sampling(data)
    Labels1, Labels2 = add_labels(Samples)
    if shuffle:
        Samples, Labels1, Labels2 = data_labels_shuffle(Samples, Labels1, Labels2)
    Train_x, Train_y1, Train_y2, Valid_x, Valid_y1, Valid_y2, Test_x, Test_y1, Test_y2 = split_data(Samples,
                                                                                                    Labels1,
                                                                                                    Labels2)
    return Train_x, Train_y1, Train_y2, Valid_x, Valid_y1, Valid_y2, Test_x, Test_y1, Test_y2


if __name__ == '__main__':
    file_path = hust_L1[0], hust_L1[1]
    train_x, train_y1, train_y2, valid_x, valid_y1, valid_y2, \
    test_x, test_y1, test_y2 = HUST_preprocess(file_path, sample_num=200, split_rate=(60, 40, 80), label1=0, label2=0)
    print(train_x.shape, train_y1.shape, train_y2.shape)
    print(valid_x.shape, valid_y1.shape, valid_y2.shape)
    print(test_x.shape, test_y1.shape, test_y2.shape)

